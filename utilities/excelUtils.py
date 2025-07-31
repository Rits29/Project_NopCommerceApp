import openpyxl


class ReadWriteExcel:
    @staticmethod
    def get_row_count(path, sheet_name):
        workbook = openpyxl.load_workbook("TestData/Login_TestData.xlsx")
        sheet = workbook["Sheet1"]
        return sheet.max_row
    @staticmethod
    def get_col_count(path, sheet_name):
        workbook = openpyxl.load_workbook("TestData/Login_TestData.xlsx")
        sheet = workbook["Sheet1"]
        return sheet.max_column
    @staticmethod
    def read_data(path, sheet_name, row, col):
        workbook = openpyxl.load_workbook("TestData/Login_TestData.xlsx")
        sheet = workbook["Sheet1"]
        return sheet.cell(row=row, column=col).value
    @staticmethod
    def write_data(rpath, sheet_name, row, col, value):
        workbook = openpyxl.load_workbook("TestData/Login_TestData.xlsx")
        sheet = workbook["Sheet1"]
        sheet.cell(row=row, column=col).value = value
        workbook.save("TestData/Login_TestData.xlsx")